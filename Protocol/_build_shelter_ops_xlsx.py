# -*- coding: utf-8 -*-
"""Smart Shelter 현장 운영·배포 구성 엑셀 생성. python _build_shelter_ops_xlsx.py"""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "Shelter_Ops_Deploy.xlsx"
OUT_V1 = Path(__file__).parent / "Shelter_Ops_Deploy_v1.0.xlsx"
BUILT = date.today().isoformat()
REV = "1.3"

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
FILL_SECTION = PatternFill("solid", fgColor="E2EFDA")
FILL_GRAY = PatternFill("solid", fgColor="F2F2F2")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_BLUE = PatternFill("solid", fgColor="D9E1F2")
FILL_ORANGE = PatternFill("solid", fgColor="FCE4D6")
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")

FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
FONT_HDR = Font(name="맑은 고딕", size=10, bold=True)
FONT_BODY = Font(name="맑은 고딕", size=9)
FONT_DIAG = Font(name="Consolas", size=9)


def style_range(ws, r1, c1, r2, c2, fill=None, font=None, align=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if align:
                cell.alignment = align


def set_row(ws, row, values, fill=FILL_WHITE, font=FONT_BODY, align=LEFT):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row, c, v)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = BORDER


def title_row(ws, row, text, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row, 1, text).font = FONT_TITLE
    ws.cell(row, 1).fill = FILL_TITLE
    ws.cell(row, 1).alignment = CENTER
    style_range(ws, row, 1, row, cols, fill=FILL_TITLE, font=FONT_TITLE, align=CENTER)
    ws.row_dimensions[row].height = 28


def section_row(ws, row, text, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row, 1, text).font = FONT_HDR
    ws.cell(row, 1).fill = FILL_SECTION
    style_range(ws, row, 1, row, cols, fill=FILL_SECTION, font=FONT_HDR, align=LEFT)


def build_overview(wb):
    ws = wb.active
    ws.title = "개요"
    ws.sheet_view.showGridLines = False
    title_row(ws, 1, "Smart Shelter 현장 운영·CMS 배포 구성 (쉘터 내부 로컬 관제)")
    rows = [
        ("문서 개정", f"rev {REV}"),
        ("생성일", BUILT),
        ("대상", "버스 쉼터 내부 — 관제 PC + 제어보드 + 공무원 접속"),
        ("핵심", "MQTT(1883) + CMS 웹(3000) — 공무원은 브라우저만 사용"),
        ("배포 패키지", "Shelter_CMS_Deploy/ (소스 코드 미포함)"),
        ("설치 가이드", "Shelter_CMS_Deploy/docs/설치_가이드.md"),
        ("운영 가이드", "Shelter_CMS_Deploy/docs/운영_가이드_공무원.md"),
        ("제어보드 최초설정", "Protocol/Shelter_Installation_Guide.md (PC 직결 192.168.0.100)"),
        ("전원보드 ID", "DIP 4bit(0~15) — CMS/MQTT tele/pb의 b_id 확인 후 제어 (고정값 없음)"),
        ("출근 스케줄", "CMS set_ac / set_ap 일괄 명령 (AC 냉방24°C·AP AUTO 2단)"),
        ("PB 제어 전제", "tele/pb로 b_id 학습 후 cmnd — 미학습 시 PB 스케줄·수동 제어 생략 (v1.3)"),
    ]
    r = 3
    set_row(ws, r, ["항목", "내용"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    for a, b in rows:
        set_row(ws, r, [a, b])
        r += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70


def build_diagram(wb):
    ws = wb.create_sheet("네트워크 구성도")
    ws.sheet_view.showGridLines = False
    title_row(ws, 1, "쉘터 내부 LAN 네트워크 구성 (도식)", 8)

    diagram = """
┌─────────────────────────────────────────────────────────────────────────────┐
│                        【 쉘터 내부 LAN 】  (예: 192.168.0.0/24)              │
│                                                                             │
│   ┌──────────────┐         ┌──────────────┐         ┌──────────────┐       │
│   │  관제 PC     │         │   스위치 /   │         │  제어보드    │       │
│   │  (미니 PC)   │◄───────►│   공유기     │◄───────►│  Smart       │       │
│   │              │  유선   │   (선택 WiFi)│  유선   │  Shelter     │       │
│   └──────┬───────┘         └──────┬───────┘         └──────┬───────┘       │
│          │                        │                        │               │
│   Mosquitto :1883                 │                 MQTT ──┘               │
│   CMS 웹    :3000                 │                 (브로커=관제PC IP)       │
│   Docker 상시가동                 │                                          │
│                                   │                                          │
│                          ┌────────┴────────┐                                 │
│                          │  공무원 노트북   │                                 │
│                          │  Chrome / Edge  │                                 │
│                          └────────┬────────┘                                 │
│                                   │                                          │
│                    http://관제PC_IP:3000  (CMS 조작·모니터링)                 │
│                    ※ Mosquitto 클라이언트 설치 불필요                         │
└─────────────────────────────────────────────────────────────────────────────┘

【 데이터 흐름 】
  제어보드 ──MQTT 1883──► 관제 PC(Mosquitto) ◄──구독── CMS(Node.js)
  공무원 PC ──HTTP 3000──► CMS 웹 UI ──Socket.IO──► CMS ──MQTT──► 제어보드

【 포트 】
  1883  MQTT (제어보드 ↔ 브로커)     |  3000  CMS 웹 (공무원 브라우저)
"""
    ws.merge_cells("A3:H28")
    cell = ws["A3"]
    cell.value = diagram.strip()
    cell.font = FONT_DIAG
    cell.alignment = LEFT
    cell.fill = FILL_BLUE
    style_range(ws, 3, 1, 28, 8, fill=FILL_BLUE, font=FONT_DIAG, align=LEFT)
    ws.row_dimensions[3].height = 420

    r = 30
    section_row(ws, r, "■ 역할 요약", 8)
    r += 1
    set_row(ws, r, ["구성요소", "설치 SW", "접속 방법", "비고"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    data = [
        ("관제 PC (쉘터 내)", "Docker + Shelter_CMS_Deploy", "24h 상시 가동", "install.bat 1회"),
        ("제어보드", "펌웨어 (Flash)", "MQTT → 관제 PC IP:1883", "최초: 직결 HTTP 설정"),
        ("공무원 PC", "브라우저만", "http://관제IP:3000", "별도 프로그램 불필요"),
    ]
    for row in data:
        set_row(ws, r, row)
        r += 1
    for c, w in zip("ABCDEFGH", [14, 22, 28, 24, 10, 10, 10, 10]):
        ws.column_dimensions[c].width = w


def build_roles(wb):
    ws = wb.create_sheet("역할·설치")
    ws.sheet_view.showGridLines = False
    title_row(ws, 1, "장비별 역할 및 설치 요구사항", 5)
    r = 3
    set_row(ws, r, ["위치", "장비", "설치·설정", "용도", "공무원 PC 필요?"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    rows = [
        ("쉘터 내부", "관제 PC (NUC/미니PC)", "Docker Desktop + install.bat", "Mosquitto + CMS 서버", "아니오 (서버 자체)"),
        ("쉘터 내부", "제어보드", "펌웨어 + Flash 네트워크 설정", "에어컨·센서·릴레이 제어", "아니오"),
        ("쉘터 내부", "전원보드 (8ch)", "DIP 4bit ID(0~15) — RS485", "전력 ON/OFF·전류 측정", "아니오"),
        ("쉘터 내부", "스위치/공유기", "(선택) Wi-Fi AP", "LAN 묶기·공무원 무선 접속", "아니오"),
        ("현장 방문", "공무원 노트북", "Chrome 또는 Edge", "CMS 웹으로 조작·점검", "예 — 브라우저만"),
        ("구청 사무실", "원격 PC", "VPN 등 (별도 설계)", "현재 패키지는 쉘터 LAN 전용", "VPN 후 브라우저"),
    ]
    for row in rows:
        set_row(ws, r, row)
        r += 1

    r += 1
    section_row(ws, r, "■ 공무원 PC — 필요 / 불필요", 5)
    r += 1
    set_row(ws, r, ["구분", "항목", "설명", "", ""], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    need = [
        ("필요", "Chrome / Edge", "CMS 접속"),
        ("필요", "쉘터 LAN 연결", "Wi-Fi 또는 유선"),
        ("필요", "CMS 주소 북마크", "http://192.168.0.107:3000 (예)"),
        ("불필요", "Mosquitto 클라이언트", "MQTT는 제어보드↔서버만"),
        ("불필요", "Node.js / Docker", "관제 PC에만 설치"),
        ("불필요", "별도 PC용 관제 SW", "CMS 웹이 UI"),
    ]
    for row in need:
        set_row(ws, r, list(row) + ["", ""])
        r += 1
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18


def build_ip_plan(wb):
    ws = wb.create_sheet("IP·포트")
    ws.sheet_view.showGridLines = False
    title_row(ws, 1, "IP 주소 · 포트 계획 (예시)", 6)
    r = 3
    set_row(ws, r, ["구분", "IP / 포트", "서브넷", "설정 방법", "비고"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    ips = [
        ("관제 PC (서버)", "192.168.0.107", "255.255.255.0", "Windows 고정 IP (STATIC 권장)", "브로커 IP = 이 주소"),
        ("제어보드 (현장)", "DHCP 또는 .50", "255.255.255.0", "직결 HTTP 최초 설정", "Flash 저장"),
        ("공무원 PC", "192.168.0.x (DHCP)", "동일 LAN", "자동 또는 고정", "서버와 같은 대역"),
        ("직결 설정 (1회)", "보드 .100 / PC .10", "255.255.255.0", "설치 업체 PC 직결", "http://192.168.0.100"),
        ("MQTT 브로커", "포트 1883", "—", "관제 PC", "제어보드 연결"),
        ("CMS 웹", "포트 3000", "—", "관제 PC", "공무원 브라우저"),
    ]
    for row in ips:
        set_row(ws, r, row)
        r += 1
    r += 1
    section_row(ws, r, "■ Windows 방화벽 (관제 PC)", 6)
    r += 1
    set_row(ws, r, ["포트", "프로토콜", "용도", "방향", "비고"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    for row in [("1883", "TCP", "MQTT", "인바운드 허용", "제어보드"), ("3000", "TCP", "CMS HTTP", "인바운드 허용", "공무원 LAN")]:
        set_row(ws, r, row)
        r += 1
    for c, w in zip("ABCDEF", [16, 18, 14, 22, 28, 20]):
        ws.column_dimensions[c].width = w


def build_workflow(wb):
    ws = wb.create_sheet("운영 흐름")
    ws.sheet_view.showGridLines = False
    title_row(ws, 1, "설치 · 일상 운영 흐름", 5)
    r = 3
    section_row(ws, r, "■ 최초 설치 (설치 업체 · 1회)", 5)
    r += 1
    set_row(ws, r, ["#", "단계", "작업", "담당", "비고"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    install = [
        ("1", "관제 PC", "Shelter_CMS_Deploy → install.bat", "설치 업체", "Docker 사전 설치"),
        ("2", "서버 IP", "고정 IP + .env SERVER_LAN_IP", "설치 업체", "예: 192.168.0.107"),
        ("3", "제어보드", "PC 직결 → http://192.168.0.100", "설치 업체", "DHCP + 브로커 IP"),
        ("4", "LAN 연결", "PC 직결 해제 → 스위치 LAN", "설치 업체", "★ 케이블 교체 필수"),
        ("5", "확인", "CMS http://서버IP:3000 → ONLINE", "설치 업체", "UID 확인"),
    ]
    for row in install:
        set_row(ws, r, row)
        r += 1
    r += 1
    section_row(ws, r, "■ 일상 점검 (공무원)", 5)
    r += 1
    set_row(ws, r, ["#", "단계", "작업", "담당", "비고"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    daily = [
        ("1", "LAN 접속", "쉘터 Wi-Fi 또는 유선", "공무원", ""),
        ("2", "CMS 접속", "http://관제IP:3000 로그인", "공무원", "기본 smart/smart 변경 권장"),
        ("3", "상태 확인", "ONLINE · 센서 · 전력", "공무원", ""),
        ("4", "조작", "에어컨·릴레이·팬 등 CMS 메뉴", "공무원", "브라우저만"),
    ]
    for row in daily:
        set_row(ws, r, row)
        r += 1
    for c, w in zip("ABCDE", [6, 14, 42, 12, 24]):
        ws.column_dimensions[c].width = w


def build_checklist(wb):
    ws = wb.create_sheet("체크리스트")
    ws.sheet_view.showGridLines = False
    title_row(ws, 1, "납품 · 인수 체크리스트", 4)
    r = 3
    set_row(ws, r, ["□", "항목", "확인 방법", "비고"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    items = [
        ("", "관제 PC Docker + install.bat 완료", "http://localhost:3000", ""),
        ("", "서버 PC 고정 IP 설정", "ipconfig", ".env 일치"),
        ("", "방화벽 1883·3000 허용", "원격 PC에서 접속", ""),
        ("", "제어보드 Flash: 브로커=서버 IP", "CMS 유지보수 또는 직결", ""),
        ("", "CMS 장치 ONLINE", "dev/tele/state", "UID 일치"),
        ("", "전원보드 b_id 확인", "dev/tele/pb 또는 CMS 전력 화면", "DIP 4bit 실측(0~15) — cmnd에 동일 ID"),
        ("", "공무원 PC CMS 접속", "http://서버IP:3000", "브라우저만"),
        ("", "CMS 로그인 비밀번호 변경", "현장 정책", "기본 smart/smart"),
        ("", "(선택) Wi-Fi SSID 안내", "공무원 전달", ""),
    ]
    for row in items:
        set_row(ws, r, row)
        r += 1
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 24


def build_faq(wb):
    ws = wb.create_sheet("FAQ")
    ws.sheet_view.showGridLines = False
    title_row(ws, 1, "자주 묻는 질문", 3)
    r = 3
    set_row(ws, r, ["질문", "답변", "참고"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    faqs = [
        ("공무원 PC에 MQTT 프로그램 필요?", "아니오. CMS 웹(브라우저)만 사용.", "포트 3000"),
        ("별도 클라이언트 앱 필요?", "아니오. Web_server_mosquitto CMS가 UI.", ""),
        ("구청에서 원격 접속?", "현재는 쉘터 LAN 전용. VPN 별도 설계.", ""),
        ("서버 IP 바꾸면?", "제어보드 브로커 IP도 재설정 필요.", "직결 또는 CMS 유지보수"),
        ("CMS 안 열림", "Docker 실행·install.bat·방화벽 3000 확인.", "설치 가이드"),
        ("장치 OFFLINE", "케이블·브로커 IP·Mosquitto 1883 확인.", ""),
        ("PB b_id가 1이 아님?", "정상. 전원보드 DIP 4bit(0~15) 실측값. tele/pb의 b_id를 cmnd에 사용.", "Shelter_Protocol §5.0"),
        ("전력 제어 안 됨", "cmnd/pb의 b_id가 tele과 불일치하면 result:0.", "get_pb로 b_id 확인"),
        ("출근 후 AC/AP만 켜짐?", "v1.2+ CMS는 set_ac/set_ap 사용. 구버전은 pwr만 적용될 수 있음.", "펌웨어·CMS 동시 갱신"),
    ]
    for row in faqs:
        set_row(ws, r, row)
        ws.row_dimensions[r].height = 36
        r += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 20


def main():
    wb = Workbook()
    build_overview(wb)
    build_diagram(wb)
    build_roles(wb)
    build_ip_plan(wb)
    build_workflow(wb)
    build_checklist(wb)
    build_faq(wb)
    wb.save(OUT)
    wb.save(OUT_V1)
    print(f"Saved: {OUT}")
    print(f"Saved: {OUT_V1}")


if __name__ == "__main__":
    main()
