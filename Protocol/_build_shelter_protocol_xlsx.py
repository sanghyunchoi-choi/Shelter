# -*- coding: utf-8 -*-
"""Generate Shelter_Protocol.xlsx (서버 업체 전달용). 수정 후: python _build_shelter_protocol_xlsx.py"""
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "Shelter_Protocol.xlsx"
OUT_V1 = Path(__file__).parent / "Shelter_Protocol_v1.0.xlsx"

# 규격 개정 — 프로토콜 변경 시 숫자 올리고 스크립트 재실행
SHELTER_PROTOCOL_REV = "2.5"
SHELTER_PROTOCOL_BUILT = date.today().isoformat()

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
FILL_DEV2SRV = PatternFill("solid", fgColor="D9E1F2")
FILL_SRV2DEV = PatternFill("solid", fgColor="FCE4D6")
FILL_SECTION = PatternFill("solid", fgColor="E2EFDA")
FILL_GRAY = PatternFill("solid", fgColor="F2F2F2")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
FONT_HDR = Font(name="맑은 고딕", size=10, bold=True)
FONT_BODY = Font(name="맑은 고딕", size=9)
FONT_MONO = Font(name="Consolas", size=9)


def style_range(ws, r1, c1, r2, c2, fill=None, font=None, align=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if align:
                cell.alignment = align


def set_row(ws, row, values, fill=FILL_WHITE, font=FONT_BODY, align=LEFT):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row, c, v)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = BORDER


def build_overview(wb):
    ws = wb.active
    ws.title = "개요"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "Smart Shelter MQTT 프로토콜 규격서 (MOS / Mosquitto)"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    info = [
        ("규격 개정", f"rev {SHELTER_PROTOCOL_REV}"),
        ("엑셀 생성일", SHELTER_PROTOCOL_BUILT),
        ("대상", "MOS_version (Mosquitto 1883, TLS 제외)"),
        ("주요 변경", "v2.5: DEV_RESET 소프트 리부트 · PB 릴레이 200ms · stat/pb conn 규칙 · CMS PB b_id 게이트"),
        ("PB b_id", "전원보드 DIP 4bit → RS485 byte[1] → 제어보드 학습 → MQTT tele/stat (고정 1 아님)"),
        ("펌웨어 설정", "config.h = Flash empty 시 fallback · 실제 설정은 직결 HTTP / Flash"),
        ("장치 식별", "dev/tele/state ONLINE 시 uid (STM32 96-bit UID)"),
        ("LAN (현장)", "Flash 저장 DHCP(권장) 또는 STATIC · 직결 HTTP로 최초 설정"),
        ("브로커 IP", "직결 HTTP 또는 Flash · CMS 유지보수 / set_broker (원격)"),
        ("직결 프로비저닝", "MQTT 90s 실패 → http://192.168.0.100 · PC=192.168.0.10 · HTML 내장"),
        ("설치 가이드", "Protocol/Shelter_Installation_Guide.md (설치 업체용)"),
        ("참고 MD", "Protocol/Shelter_Protocol.md (MQTT 상세·미구현 항목)"),
        ("자동 보고", "10분 주기 (재연결·명령 후 즉시 1회 + 타이머 리셋)"),
        ("전원보드", "RS485 ~0.8s 폴링 → MQTT tele 10s"),
        ("토픽 prefix", "dev/tele/*  dev/stat/*  dev/cmnd/*"),
        ("페이로드", "JSON (UTF-8), QoS 0"),
    ]
    r = 3
    for k, v in info:
        ws.cell(r, 1, k).font = FONT_HDR
        ws.cell(r, 1).fill = FILL_GRAY
        ws.cell(r, 1).border = BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(r, 2, v).font = FONT_BODY
        ws.cell(r, 2).fill = FILL_WHITE
        ws.cell(r, 2).border = BORDER
        ws.cell(r, 2).alignment = LEFT
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r, 1, "■ 통신 방향 (토픽 구조)")
    ws.cell(r, 1).font = FONT_HDR
    ws.cell(r, 1).fill = FILL_SECTION
    ws.cell(r, 1).alignment = LEFT
    style_range(ws, r, 1, r, 8, fill=FILL_SECTION, font=FONT_HDR, align=LEFT)
    r += 1

    flow_hdr = ["방향", "Prefix", "토픽 예시", "발행 주체", "구독 주체", "설명", "", ""]
    set_row(ws, r, flow_hdr, FILL_DEV2SRV, FONT_HDR, CENTER)
    r += 1
    flows = [
        ("장치 → 서버", "dev/tele/", "dev/tele/dust", "제어보드", "웹서버/브로커", "Telemetry · 10분 자동 보고"),
        ("장치 → 서버", "dev/stat/", "dev/stat/pb", "제어보드", "웹서버/브로커", "명령 처리 결과 (result 포함)"),
        ("서버 → 장치", "dev/cmnd/", "dev/cmnd/pb", "웹서버", "제어보드", "제어 명령 (dev/cmnd/# 구독)"),
    ]
    for row in flows:
        set_row(ws, r, list(row) + ["", ""])
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r, 1, "■ 장치 목록 (7종 + 시스템)")
    ws.cell(r, 1).font = FONT_HDR
    ws.cell(r, 1).fill = FILL_SECTION
    style_range(ws, r, 1, r, 8, fill=FILL_SECTION, font=FONT_HDR, align=LEFT)
    r += 1

    dev_hdr = ["#", "장치", "약어", "tele", "stat", "cmnd", "비고", ""]
    set_row(ws, r, dev_hdr, FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    devices = [
        ("1", "장비 상태", "state", "O", "-", "-", "LWT OFFLINE"),
        ("2", "먼지센서", "dust", "O", "-", "-", "tele 전용"),
        ("3", "내부 온습도", "th_in", "O", "-", "-", "tele 전용"),
        ("4", "외부 온습도", "th_out", "O", "-", "-", "tele 전용"),
        ("5", "자동문 릴레이", "ad", "O", "O", "O", "15채널"),
        ("6", "내부 팬", "fan", "O", "O", "O", "AUTO/MANUAL · rpm · 타코 conn"),
        ("7", "전원보드", "pb", "O", "O", "O", "8채널 · MOS/TLS 공용"),
        ("8", "공기청정기", "ap", "O", "O", "O", "Himpel"),
        ("9", "에어컨(LG)", "ac", "O", "O", "O", "Modbus"),
        ("10", "시스템", "dev", "-", "O", "O", "RESET / ALL_DATA / NET"),
        ("11", "외부 입력 4ch", "input", "O", "O", "O", "PD4~PD7 · 보드 부착"),
    ]
    for row in devices:
        set_row(ws, r, list(row) + [""])
        r += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 8


def build_protocol_sheet(wb):
    ws = wb.create_sheet("프로토콜 정의")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "MQTT 제어 프로토콜 — 장치별 통신 규격"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Direction group headers (row 2-3)
    ws.merge_cells("A2:A3")
    ws["A2"] = "장치"
    ws.merge_cells("B2:E2")
    ws["B2"] = "◀ 장치 → 서버 (Publish)"
    ws.merge_cells("F2:H2")
    ws["F2"] = "서버 → 장치 (Subscribe) ▶"
    ws.merge_cells("I2:J2")
    ws["I2"] = "비고"

    sub_hdr = [
        None,  # A3 merged with A2
        "구분", "토픽", "연결 시 Payload", "미연결 Payload",
        "토픽", "명령 Payload 예시", "설명",
        "필드 설명", "주기/특이사항",
    ]
    for c, h in enumerate(sub_hdr, 1):
        if h is None:
            continue
        cell = ws.cell(3, c, h)
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = BORDER
        if 2 <= c <= 5:
            cell.fill = FILL_DEV2SRV
        elif 6 <= c <= 8:
            cell.fill = FILL_SRV2DEV
        else:
            cell.fill = FILL_GRAY

    rows = [
        # device, d2s_type, d2s_topic, d2s_conn, d2s_disc,
        # s2d_topic, s2d_cmd, s2d_desc, fields, notes
        [
            "장비\n상태",
            "tele\n(LWT)",
            "dev/tele/state",
            '{"status":"ONLINE","uid":"A1B2C3D4E5F6789012345678","conn":1}',
            '{"status":"OFFLINE"}',
            "(없음)",
            "(없음)",
            "명령 없음",
            "status: ONLINE/OFFLINE\nuid: STM32 96-bit UID (24 hex)\nconn: 0/1",
            "LWT retained=1\n비정상 종료 시 OFFLINE",
        ],
        [
            "먼지\n센서",
            "tele",
            "dev/tele/dust",
            '{"pm1_0":15,"pm2_5":24,"pm10":24,"conn":1}',
            '{"dust":0,"conn":0}',
            "(없음)",
            "(없음)",
            "tele 전용",
            "pm1_0/pm2_5/pm10 [㎍/㎥]",
            "10분 주기",
        ],
        [
            "내부\n온습도",
            "tele",
            "dev/tele/th_in",
            '{"temp":25.3,"humi":42.1,"conn":1}',
            '{"thindoor":0,"conn":0}',
            "(없음)",
            "(없음)",
            "tele 전용",
            "temp [°C], humi [%RH]",
            "10분 주기",
        ],
        [
            "외부\n온습도",
            "tele",
            "dev/tele/th_out",
            '{"temp":12.5,"humi":35.0,"conn":1}',
            '{"thoutdoor":0,"conn":0}',
            "(없음)",
            "(없음)",
            "tele 전용",
            "temp [°C], humi [%RH]",
            "10분 주기",
        ],
        [
            "자동문\n릴레이\n(15ch)",
            "tele",
            "dev/tele/ad",
            '{"relays":[1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],"conn":1}',
            "(보드 부착 —\n미연결 없음)",
            "dev/cmnd/ad",
            '① {"ch":1,"relay_1":1}\n② {"set_ad":"100000000000000"}\n③ {"get_ad":"state"}',
            "개별 / 전체(15자리) / 조회",
            "relays: 15ch [0/1]\nch: 1~15\nrelay_N: 0/1",
            "전체 set 시 500ms 간격\n순차 제어",
        ],
        [
            "자동문\n릴레이",
            "stat\n(응답)",
            "dev/stat/ad",
            '{"relays":[1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],"conn":1,"result":1}',
            '{"relays":[...],"conn":1,"result":0}',
            "dev/cmnd/ad",
            "(위 cmnd 참조)",
            "cmnd 처리 후 stat 발행",
            "result: 1=성공 0=실패",
            "명령 후 10분 타이머 리셋",
        ],
        [
            "내부\n팬",
            "tele",
            "dev/tele/fan",
            '{"mode":"MANUAL","duty":50,"rpm":1234,"conn":1}',
            '{"fan":0,"rpm":0,"conn":0}',
            "dev/cmnd/fan",
            '① {"set_fan":{"mode":"AUTO","duty":0}}\n② {"set_fan":{"mode":"MANUAL","duty":50}}\n③ {"get_fan":"state"}',
            "AUTO / MANUAL / 조회",
            "mode: AUTO|MANUAL\nduty: 0~100 [%]\nrpm: 타코 RPM",
            "conn: duty≥10%→7s 타코\nduty<10%→60s 이력\nAUTO: 내부온도",
        ],
        [
            "내부\n팬",
            "stat\n(응답)",
            "dev/stat/fan",
            '{"mode":"MANUAL","duty":50,"rpm":1234,"conn":1,"result":1}',
            '{"fan":0,"rpm":0,"conn":0,"result":1}',
            "dev/cmnd/fan",
            "(위 cmnd 참조)",
            "cmnd 처리 후 stat 발행",
            "result: 1/0",
            "10분 타이머 리셋",
        ],
        [
            "전원\n보드\n(8ch)",
            "tele",
            "dev/tele/pb",
            '{\n  "b_id":3,\n  "pb":[\n    {"ch":1,"sw":1,"c":1.14,"w":250.8,"v":220},\n    {"ch":2,"sw":1,"c":1.04,"w":228.5,"v":220},\n    ... ch3~8 ...\n  ],\n  "conn":1\n}',
            '{"pb":0,"conn":0}',
            "dev/cmnd/pb",
            '① {"b_id":N,"ch":1,"sw":1}\n② {"b_id":N,"set_pb":"11001101"}\n③ {"b_id":N,"get_pb":"state"}\n(N=tele/get_pb 실측 b_id)',
            "개별 / 8자리 일괄 / 조회",
            "b_id: DIP 4bit(0~15)\nPB RS485 byte[1]\n제어보드 학습값\nch:1~8 sw:0/1\nc[A] w[W] v[V]",
            "RS485 ~0.8s 폴링\nMQTT tele 10s\nconn: HAL_OK / 10회 fail",
        ],
        [
            "전원\n보드",
            "stat\n(응답)",
            "dev/stat/pb",
            '(tele 동일 +\n"result":1)',
            '{"b_id":N,"pb":0,"conn":1,"result":0}',
            "dev/cmnd/pb",
            "(위 cmnd 참조)",
            "cmnd 처리 후 stat 발행",
            "result: 1/0",
            "10분 타이머 리셋",
        ],
        [
            "공기\n청정기\n(Himpel)",
            "tele",
            "dev/tele/ap",
            '{"pwr":1,"mode":"AUTO","spd":3,"uv":1,\n"co2":450,"pm25":15,"pm10":20,"pm1_0":10,\n"temp":23.5,"temp_out":18.2,"humi":45,"tvoc":120,"filter":100,"conn":1}',
            '{"ap":0,"conn":0}',
            "dev/cmnd/ap",
            '개별·복합:\n{"pwr":0} {"mode":"AUTO"}\n{"spd":4} {"uv":1}\n{"pwr":1,"mode":"AUTO","spd":2}\n\n일괄:\n{"set_ap":{...}}\n\n조회:\n{"get_ap":"state"}',
            "개별·복합 / set_ap / get_ap",
            "pwr 0/1, spd 0~4\nmode: AUTO/MANUAL/CLEAN/VENT/BYPASS/HEATER\ntemp, temp_out, tvoc",
            "Himpel RS485\n(Modbus 아님)\nconn: RS485 OK",
        ],
        [
            "공기\n청정기",
            "stat\n(응답)",
            "dev/stat/ap",
            '(tele +\n"result":1)',
            '{"ap":0,"conn":0,"result":0}',
            "dev/cmnd/ap",
            "(위 cmnd 참조)",
            "cmnd 처리 후 stat 발행",
            "result: 1/0",
            "10분 타이머 리셋",
        ],
        [
            "에어컨\n(LG)",
            "tele",
            "dev/tele/ac",
            '{"pwr":1,"mode":"COOL","temp":24,"curr":26.0,"spd":3,"err":0,"conn":1}',
            '{"ac":0,"conn":0}',
            "dev/cmnd/ac",
            '개별·복합:\n{"pwr":1} {"mode":"COOL"}\n{"temp":26} {"spd":3}\n{"pwr":1,"mode":"COOL","temp":24,"spd":2}\n\n일괄:\n{"set_ac":{...}}\n\n조회:\n{"get_ac":"state"}',
            "개별·복합 / set_ac / get_ac",
            "mode: COOL/HEAT/FAN/DRY/AUTO\ntemp 16~30, spd 1~6",
            "LG Modbus RTU\nconn: Discrete\nADDR_AC_CONN bit",
        ],
        [
            "에어컨",
            "stat\n(응답)",
            "dev/stat/ac",
            '(tele +\n"result":1)',
            '{"ac":0,"conn":0,"result":0}',
            "dev/cmnd/ac",
            "(위 cmnd 참조)",
            "cmnd 처리 후 stat 발행",
            "result: 1/0",
            "10분 타이머 리셋",
        ],
        [
            "외부\n입력\n(4ch)",
            "tele",
            "dev/tele/input",
            '{"in":[1,0,0,1],"count":[12,0,0,3],"conn":1}',
            "(보드 부착 —\n미연결 없음)",
            "dev/cmnd/input",
            '① {"get_input":"state"}\n② {"reset_count":1}',
            "조회 / 카운트 초기화",
            "in: PD4~PD7 [0/1]\ncount: 상승엣지 누적\nconn: 1 (보드 부착)",
            "상태변경 즉시 dev/tele/input\n(서버 폴링 없음)",
        ],
        [
            "외부\n입력",
            "stat\n(응답)",
            "dev/stat/input",
            '{"in":[1,0,0,1],"count":[12,0,0,3],"conn":1,"result":1}',
            '{"in":[0,0,0,0],"count":[0,0,0,0],"conn":1,"result":0}',
            "dev/cmnd/input",
            "(위 cmnd 참조)",
            "get_input / reset_count 응답",
            "result: 1/0",
            "연결 표시 불필요 (보드 GPIO)",
        ],
        [
            "시스템\n제어",
            "stat\n(응답)",
            "dev/stat/dev",
            '{"device":"ALL_DATA","conn":1,"result":1}\n+ tele 8종 일괄 발행',
            '{"device":"DEV_RESET","conn":0,"result":1}',
            "dev/cmnd/dev",
            '① {"reset":"DEV_RESET"}\n   → 500ms 후 소프트 리부트\n     (Flash·네트워크 유지)\n\n② {"data":"ALL_DATA"}\n   → stat ACK +\n     tele 전체 즉시 발행\n\n③ {"set_net":{"ip":[...],"sn":[...],"gw":[...]}}\n   → STATIC IP Flash 저장\n\n④ {"set_broker":{"ip":[...],"port":1883}}\n   → 브로커 Flash 저장\n\n⑤ {"set_net_reset":1}\n   → Flash DHCP 공장 초기화\n\n⑥ 직결 HTTP (v2.1)\n   → 192.168.0.100:80\n     PC 192.168.0.10\n\n⑦ {"clear_net":1} (예정)\n\n⑧ {"clear_broker":1} (예정)',
            "리셋 / 전체 / NET / 브로커",
            "ALL_DATA 발행:\nth_in, th_out, dust,\nad, fan, ac, pb, ap, input",
            "clear_* : 자체 8ch 스위치\n별도 제어보드 IP/브로커\n삭제 (펌웨어 예정)",
        ],
    ]

    r = 4
    prev_device = None
    merge_start = None
    for row in rows:
        device = row[0]
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.font = FONT_MONO if c in (4, 5, 7) else FONT_BODY
            cell.alignment = LEFT
            cell.border = BORDER
            if 2 <= c <= 5:
                cell.fill = FILL_DEV2SRV if c > 1 else FILL_WHITE
            elif 6 <= c <= 8:
                cell.fill = FILL_SRV2DEV if c > 5 else FILL_WHITE
            else:
                cell.fill = FILL_WHITE
        ws.row_dimensions[r].height = max(60, 15 * str(row[6]).count("\n"))

        # merge device name column for tele+stat pairs
        base = device.split("\n")[0]
        if prev_device and base == prev_device.split("\n")[0]:
            pass
        else:
            if merge_start and merge_start < r - 1:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=r - 1, end_column=1)
                ws.cell(merge_start, 1).alignment = CENTER
            merge_start = r
        prev_device = device
        r += 1
    if merge_start and merge_start < r - 1:
        ws.merge_cells(start_row=merge_start, start_column=1, end_row=r - 1, end_column=1)
        ws.cell(merge_start, 1).alignment = CENTER

    # style direction headers row 2
    style_range(ws, 2, 1, 3, 1, fill=FILL_GRAY, font=FONT_HDR, align=CENTER)
    style_range(ws, 2, 2, 2, 5, fill=FILL_DEV2SRV, font=FONT_HDR, align=CENTER)
    style_range(ws, 2, 6, 2, 8, fill=FILL_SRV2DEV, font=FONT_HDR, align=CENTER)
    style_range(ws, 2, 9, 3, 10, fill=FILL_GRAY, font=FONT_HDR, align=CENTER)

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 42
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 22

    ws.freeze_panes = "B4"


def build_flow_sheet(wb):
    ws = wb.create_sheet("통신 흐름")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "장치 ↔ 서버 통신 흐름도"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER

    blocks = [
        ("A", "센서류 (tele 전용)", "장치 ──tele──▶ 서버", [
            "dev/tele/state   ONLINE / OFFLINE (LWT)",
            "dev/tele/dust    pm1_0, pm2_5, pm10",
            "dev/tele/th_in   내부 온습도",
            "dev/tele/th_out  외부 온습도",
        ], "10분 주기 자동 발행\n재연결 시 1회 즉시"),
        ("B", "제어 장비 (tele + stat + cmnd)", "", [], ""),
    ]

    r = 3
    controlled = [
        ("AD (릴레이 15ch)", "dev/tele/ad", "dev/stat/ad", "dev/cmnd/ad",
         "서버 ──cmnd──▶ 장치 ──stat──▶ 서버\n장치 ──tele──▶ 서버 (10분)", "15채널 릴레이"),
        ("Fan (팬)", "dev/tele/fan", "dev/stat/fan", "dev/cmnd/fan",
         "set_fan / get_fan", "AUTO/MANUAL · rpm · 타코 conn"),
        ("PB (전원 8ch)", "dev/tele/pb", "dev/stat/pb", "dev/cmnd/pb",
         "set_pb(8자리) / get_pb", "RS485 ~0.8s · tele 10s"),
        ("AP (공기청정)", "dev/tele/ap", "dev/stat/ap", "dev/cmnd/ap",
         "set_ap / 개별·복합 pwr,mode,spd…", "Himpel RS485"),
        ("AC (에어컨)", "dev/tele/ac", "dev/stat/ac", "dev/cmnd/ac",
         "set_ac / 개별 pwr,mode,temp…", "LG Modbus RTU"),
        ("DEV (시스템)", "(없음)", "dev/stat/dev", "dev/cmnd/dev",
         "RESET / ALL_DATA → tele 8종 일괄", "서버 최초 연결 시 ALL_DATA"),
    ]

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1, "■ A. 센서류 — 장치 → 서버 (단방향 tele)")
    ws.cell(r, 1).fill = FILL_DEV2SRV
    ws.cell(r, 1).font = FONT_HDR
    style_range(ws, r, 1, r, 6, fill=FILL_DEV2SRV, font=FONT_HDR, align=LEFT)
    r += 1
    set_row(ws, r, ["장치", "방향", "토픽", "Payload 예시", "주기", "비고"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    sensor_rows = [
        ("장비 상태", "장치→서버", "dev/tele/state", '{"status":"ONLINE","uid":"A1B2C3D4E5F6789012345678","conn":1}', "연결/해제", "LWT OFFLINE"),
        ("먼지", "장치→서버", "dev/tele/dust", '{"pm1_0":15,"pm2_5":24,"pm10":24,"conn":1}', "10분", ""),
        ("내부 온습", "장치→서버", "dev/tele/th_in", '{"temp":25.3,"humi":42.1,"conn":1}', "10분", ""),
        ("외부 온습", "장치→서버", "dev/tele/th_out", '{"temp":12.5,"humi":35.0,"conn":1}', "10분", ""),
        ("외부 입력 4ch", "장치→서버", "dev/tele/input", '{"in":[1,0,0,1],"count":[12,0,0,3],"conn":1}', "상태변경 즉시", "PD4~PD7 · 연결 표시 없음"),
    ]
    for row in sensor_rows:
        set_row(ws, r, row)
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1, "■ B. 제어 장비 — 양방향 (cmnd → stat, tele 자동 보고)")
    ws.cell(r, 1).fill = FILL_SRV2DEV
    style_range(ws, r, 1, r, 6, fill=FILL_SECTION, font=FONT_HDR, align=LEFT)
    r += 1
    set_row(ws, r, ["장치", "서버→장치 (cmnd)", "장치→서버 (stat)", "장치→서버 (tele)", "흐름", "비고"], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    for name, tele, stat, cmnd, flow, note in controlled:
        set_row(ws, r, [name, cmnd, stat, tele, flow, note])
        r += 1

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1, "■ C. ALL_DATA 시퀀스")
    ws.cell(r, 1).fill = FILL_SECTION
    style_range(ws, r, 1, r, 6, fill=FILL_SECTION, font=FONT_HDR, align=LEFT)
    r += 1
    seq = [
        ("1", "서버→장치", "dev/cmnd/dev", '{"data":"ALL_DATA"}'),
        ("2", "장치→서버", "dev/stat/dev", '{"device":"ALL_DATA","conn":1,"result":1}'),
        ("3", "장치→서버", "dev/tele/th_in ~ ap", "각 tele 토픽 즉시 1회 발행 (8종)"),
        ("4", "-", "-", "10분 타이머 전체 리셋"),
    ]
    set_row(ws, r, ["순서", "방향", "토픽", "Payload / 동작", "", ""], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    for row in seq:
        set_row(ws, r, list(row) + ["", ""])
        r += 1

    for col, w in zip("ABCDEF", [14, 14, 18, 42, 28, 22]):
        ws.column_dimensions[col].width = w


def build_rules_sheet(wb):
    ws = wb.create_sheet("동작 규칙")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = "동작 규칙 & 재연결 정책"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER

    set_row(ws, 3, ["#", "규칙", "설명", "비고"], FILL_GRAY, FONT_HDR, CENTER)
    rules = [
        ("R1", "최초 전원 인가", "각 장비 탐색 → conn 여부에 맞는 페이로드 1회 즉시 발행 → 10분 주기", "못 찾으면 conn:0"),
        ("R2", "10분 정기 보고", "tele 순차. PB: RS485 ~0.8s + MQTT 10s", "전원보드 conn"),
        ("R3", "센서 재연결", "연결 감지 → 1회 즉시 + 10분 타이머 리셋", ""),
        ("R4", "브로커 재연결", "복구 후 전체 1회 + ALL_DATA 가능", "서버 권장"),
        ("R5", "result 규칙", "성공=1+현재상태, 실패=0, 범위초과=0", "stat 전용"),
        ("R6", "LWT", "비정상 종료 → dev/tele/state OFFLINE", "retained"),
        ("R7", "명령 후 리셋", "stat 성공 시 해당 장비 10분 카운트 리셋", ""),
        ("R8", "직결 프로비저닝", "링크UP+IP없음 5s → 192.168.0.100 · 저장직후 pending=1 → LAN 2분", "양산·최초설치 표준"),
        ("R9", "설정 우선순위", "Flash > config.h fallback", "CMS=유지보수"),
        ("R10", "conn 독립 (v2.2)", "장치별 tele conn만 UI. 메인 MQTT ≠ 전체 연결", "CMS server.js"),
        ("R11", "CMS 메인 타임아웃", "5분 MQTT 없음 → state OFFLINE + 하위 conn=0", "DEVICE_TIMEOUT_MS"),
        ("R12", "장치 conn 판정", "dust 60s / fan 타코 / PB RS485×10 / AC discrete", "시트 '연결상태'"),
        ("R13", "복합 cmnd (v2.4)", "AC/AP JSON 다필드 순차 적용 · set_ac/set_ap 권장", "mqtt_handler.c"),
        ("R14", "tele vs stat", "result는 stat 전용 · tele/pb에 result 없음", "app.c"),
        ("R15", "PB b_id", "서버 b_id 하드코딩 금지 · 초기 0", "server.js"),
    ]
    r = 4
    for row in rules:
        set_row(ws, r, row)
        r += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 18


def build_provisioning_sheet(wb):
    ws = wb.create_sheet("설치·프로비저닝")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "Smart Shelter 설치·프로비저닝 (v2.1 — 설치 업체·양산)"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    ws["A2"] = "상세 절차: Protocol/Shelter_Installation_Guide.md"
    ws["A2"].font = FONT_BODY
    ws["A2"].alignment = LEFT
    style_range(ws, 2, 1, 2, 5, fill=FILL_GRAY, font=FONT_BODY, align=LEFT)

    set_row(ws, 4, ["#", "단계", "작업", "값 / URL", "비고"], FILL_GRAY, FONT_HDR, CENTER)

    steps = [
        ("1", "펌웨어", "공통 이미지 플래시", "code_mosquitto_choi", "고객별 config.h 빌드 불필요"),
        ("2", "직결", "PC ↔ 제어보드 이더넷", "케이블 1m+", "공유기 불필요"),
        ("3", "대기", "설정 모드 진입", "링크 UP + 5s", "재설정·공장 / 케이블 필수"),
        ("4", "PC IP", "Windows 수동 IP", "192.168.0.10 / 255.255.255.0", "Wi-Fi 끄기 권장"),
        ("5", "브라우저", "설정 페이지", "http://192.168.0.100", "포트 80 · XML 불필요"),
        ("6", "LAN", "모드 선택", "DHCP (권장)", "STATIC = 고정 IP만"),
        ("7", "브로커", "MQTT 입력", "관제 PC LAN IP : 1883", "Flash 저장"),
        ("8", "저장", "저장 및 재부팅", "save.cgi → pending=1", "Flash → reboot"),
        ("9", "현장", "★ PC 케이블 분리 → LAN", "공유기/스위치", "2분 DHCP 유예"),
        ("10", "확인", "CMS ONLINE", "dev/tele/state uid", "설치 완료"),
    ]
    r = 5
    for row in steps:
        set_row(ws, r, row)
        ws.row_dimensions[r].height = 36
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(r, 1, "■ IP 요약")
    ws.cell(r, 1).fill = FILL_SECTION
    ws.cell(r, 1).font = FONT_HDR
    style_range(ws, r, 1, r, 5, fill=FILL_SECTION, font=FONT_HDR, align=LEFT)
    r += 1

    ip_rows = [
        ("제어보드 (직결)", "192.168.0.100", "255.255.255.0", "프로비저닝 전용", ""),
        ("설정 PC", "192.168.0.10", "255.255.255.0", "수동 설정", ""),
        ("제어보드 (현장)", "DHCP 할당", "—", "예: 192.168.0.x", "공유기 LAN"),
        ("MQTT 브로커", "서버 LAN IP", "—", "포트 1883", "MOS 비암호화"),
    ]
    set_row(ws, r, ["구분", "IP", "서브넷", "설명", ""], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    for row in ip_rows:
        set_row(ws, r, list(row))
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(r, 1, "■ 공장 초기화 / 유지보수")
    ws.cell(r, 1).fill = FILL_SECTION
    ws.cell(r, 1).font = FONT_HDR
    style_range(ws, r, 1, r, 5, fill=FILL_SECTION, font=FONT_HDR, align=LEFT)
    r += 1

    maint = [
        ("SW8", "Flash erase → DHCP 공장값", "최초 설정 재진행"),
        ("set_net_reset", '{"set_net_reset":1}', "MQTT — ONLINE 시"),
        ("CMS 유지보수", "🔧 네트워크 메뉴", "최초 설치 X · 원격 변경용"),
    ]
    set_row(ws, r, ["방법", "명령/동작", "비고", "", ""], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    for row in maint:
        set_row(ws, r, list(row) + ["", ""])
        r += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 22


def build_conn_sheet(wb):
    """v2.2 — 장치별 conn(연결됨/오프라인) 판정 (펌웨어 + CMS)."""
    ws = wb.create_sheet("연결상태")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "conn — 연결됨 / 오프라인 판정 (rev 2.2)"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "CMS UI 배지 = 각 tele/stat의 conn 필드. 메인보드 MQTT 수신 시 하위 장치 conn 을 일괄 1로 만들지 않음."
    )
    ws["A2"].font = FONT_BODY
    ws["A2"].alignment = LEFT
    style_range(ws, 2, 1, 2, 6, fill=FILL_GRAY, font=FONT_BODY, align=LEFT)

    set_row(ws, 4, ["장치", "conn=1 (펌웨어)", "conn=0 / 미연결 payload", "CMS 표시", "타임아웃·주기", "비고"],
            FILL_GRAY, FONT_HDR, CENTER)

    rows = [
        ("state\n(메인보드)", "MQTT 브로커 연결", '{"status":"OFFLINE"} (LWT)',
         "state.conn=1 (MQTT 수신)", "5분 무수신→OFFLINE", "uid/ip/mode 포함"),
        ("dust", "UART 패킷 수신·60s 이내 갱신", '{"dust":0,"conn":0}',
         "payload.conn", "tele 10분", "UPDM010UB"),
        ("th_in", "HC-SD Modbus OK + 유효 온도", '{"thindoor":0,"conn":0}',
         "payload.conn", "읽기 3s · fail×30", "내부 온습"),
        ("th_out", "CWT-TH03S Modbus OK", '{"thoutdoor":0,"conn":0}',
         "payload.conn", "읽기 2.5s · fail×10", "외부 온습"),
        ("ad", "제어보드 릴레이 (항상 부착)", "(없음)",
         "payload.conn=1", "tele 10분", "미연결 개념 없음"),
        ("fan", "duty≥10%: 타코 7s 이내\n duty<10%: 60s 이력", '{"fan":0,"rpm":0,"conn":0}',
         "payload.conn + rpm", "conn 변경 즉시", "TIM1 CH3 타코"),
        ("pb", "RS485 UpdateAllData HAL_OK", '{"b_id":N,"pb":0,"conn":0}',
         "pb[]+conn; graph 35s stale", "tele 10s · fail×10", "전원보드 RS485"),
        ("ap", "Himpel RS485 응답 OK", '{"ap":0,"conn":0}',
         "payload.conn", "10s 무응답→0", "공기청정기"),
        ("ac", "Modbus Discrete ADDR_AC_CONN\n읽기 OK 且 bit=1", '{"ac":0,"conn":0}',
         "payload.conn", "폴링 ~1s", "LG Modbus"),
        ("input", "PD4~PD7 GPIO", "(없음)",
         "conn=1", "변경 즉시", "외부 입력 4ch"),
    ]

    r = 5
    for row in rows:
        set_row(ws, r, row)
        ws.row_dimensions[r].height = 48
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1, "■ CMS (server.js) 요약")
    ws.cell(r, 1).fill = FILL_SECTION
    ws.cell(r, 1).font = FONT_HDR
    style_range(ws, r, 1, r, 6, fill=FILL_SECTION, font=FONT_HDR, align=LEFT)
    r += 1

    cms = [
        ("normalizeConn", "tele/stat JSON conn 필드만 0/1 변환", "기본값 0"),
        ("resetDeviceTimeout", "아무 MQTT 수신 → state.conn=1, 5분 타이머", "하위 conn 일괄 변경 X"),
        ("setAllOffline", "5분 무수신 → state OFFLINE, dust/th/ad/fan/ap/ac/input conn=0", "pb 키 제외"),
        ("PB tele", "pb[] 배열 + conn → conn=1; pb:0 → conn=0 (tele에 result 없음)", "PB_TELE_STALE_MS=35s"),
        ("PB b_id", "current_b_id 초기 0 — tele 수신 후 캐시", "하드코딩 금지"),
        ("스케줄", "출근 set_ac / set_ap 일괄 명령", "server.js"),
        ("stat result:0", "해당 장치 conn=0", "명령 실패 반영"),
    ]
    set_row(ws, r, ["기능", "동작", "비고", "", "", ""], FILL_GRAY, FONT_HDR, CENTER)
    r += 1
    for row in cms:
        set_row(ws, r, list(row))
        r += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.freeze_panes = "A5"


def main():
    wb = Workbook()
    build_overview(wb)
    build_protocol_sheet(wb)
    build_flow_sheet(wb)
    build_conn_sheet(wb)
    build_rules_sheet(wb)
    build_provisioning_sheet(wb)
    wb.save(OUT)
    wb.save(OUT_V1)
    print(f"Created: {OUT}")
    print(f"Created: {OUT_V1}")


if __name__ == "__main__":
    main()
